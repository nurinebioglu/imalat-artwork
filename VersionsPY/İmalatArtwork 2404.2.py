import openpyxl.reader.excel
from customtkinter import *
from tkinter import *
from PIL import Image, ImageTk
from tkinter import filedialog
from datetime import datetime
from PIL import ImageGrab
from tkinter import Scale
from tkinter import colorchooser  # Renk seçim paletini ekleyin
import customtkinter
import sys, os
import subprocess
import shutil
import tempfile
import requests
import urllib.request
from tkinter import messagebox, simpledialog
from urllib.parse import unquote
import configparser
import ctypes
from tkinter.ttk import Combobox
import openpyxl , xlrd
from openpyxl import Workbook, load_workbook
import pathlib
from openpyxl import Workbook
from openpyxl.styles import Alignment
import re
from openpyxl.styles import Font

# Mevcut sürüm bilgisini kodlarınız arasında bir değişken olarak tutun

LOCAL_VERSION = "2404.2"
REMOTE_VERSION_URL = "https://raw.githubusercontent.com/nurinebioglu/imalat-artwork/main/İmalatArtworkMissyVersion.txt"  # Uzaktaki sürüm dosyasının URL'si
UPDATE_FILE_URL = "https://github.com/nurinebioglu/imalat-artwork/raw/main/Missy%C4%B0malatArtwork.exe"  # Güncelleme dosyasının URL'si

# Önceden tanımlı şifre

predefined_password = "nurinebioglu"


def is_admin():
    try:
        return ctypes.windll.shell32.IsUserAnAdmin()
    except:
        return False

def run_as_admin():
    if not is_admin():
        onay = messagebox.askyesno("Yönetici İzni", "Uygulamayı yönetici olarak çalıştırmak istiyor musunuz?")
        if onay:
            # Uygulamayı yönetici olarak yeniden başlat
            ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, __file__, None, 1)
            sys.exit()
    else:
        create_firewall_rule()

def create_firewall_rule():
    uygulama_dizini = os.getcwd()
    subprocess.run(["netsh", "advfirewall", "firewall", "add", "rule", "name=MyApp", "dir=in", "action=allow", f"program={uygulama_dizini}\\{__file__}", "enable=yes"])
    print("Güvenlik istisnaları başarıyla oluşturuldu.")

create_firewall_rule()

# Uygulamayı yönetici olarak çalıştır ve güvenlik istisnalarını oluştur
run_as_admin()


# .txt dosyasının URL'si
file_url = "https://raw.githubusercontent.com/nurinebioglu/imalat-artwork/main/İmalatArtworkŞifreMissy.txt"

# .txt dosyasını indirme
response = requests.get(file_url)

# .txt dosyasından şifreyi al
file_content = response.text.strip()






def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

# Global değişkenlerin tanımlanması
image_path = ""
second_image_path = ""
third_image_path = ""
image_item = None
image_img = None
second_image_item = None
second_image_img = None
third_image_item = None
third_image_img = None
frame_count = 0
textboxes = {}
frame = None  # global değişken olarak tanımlanıyor
handle = None  # global değişken olarak tutamaç ekleniyor
APP_DIRECTORY = os.getcwd()
# Resmin seçilip seçilmediğini belirleyen değişken
image_selected = False





#--------------

def check_for_update():
    try:
        response = requests.get(REMOTE_VERSION_URL)
        response.raise_for_status()
        remote_version = response.text.strip()

        if remote_version != LOCAL_VERSION:
            messagebox.showinfo("Güncelleme",
                                "Yeni bir güncelleme mevcut!\nMevcut sürüm: {}\nUzaktaki sürüm: {}".format(
                                    LOCAL_VERSION, remote_version))
            return remote_version  # Güncellenecek versiyon numarasını döndür
        else:
            messagebox.showinfo("Güncelleme", "Uygulama zaten güncel. Mevcut sürüm: {}".format(LOCAL_VERSION))
            return None  # Güncelleme gerekmediği için None döndür
    except Exception as e:
        messagebox.showerror("Hata", "Uzaktaki sürüm bilgisine ulaşılamadı: {}".format(str(e)))
        return None  # Sürüm bilgisine ulaşılamadığı için None döndür

def download_update(remote_version):
    try:
        response = requests.get(UPDATE_FILE_URL)
        response.raise_for_status()

        # Dosya adını otomatik olarak al
        filename = os.path.basename(UPDATE_FILE_URL)

        # Unicode karakterleri URL kodundan dönüştür
        filename = unquote(filename, 'utf-8')

        # Version numarasını dosya adına ekle
        filename_without_ext, ext = os.path.splitext(filename)
        versioned_filename = f"{filename_without_ext}_{remote_version}{ext}"

        # Kullanıcıya nereye kaydetmek istediğini sor
        file_path = filedialog.asksaveasfilename(
            initialdir=os.path.expanduser("~/Desktop"),
            title="Yeni güncelleme dosyasını kaydedin",
            defaultextension=".exe",
            initialfile=versioned_filename
        )

        # Eğer kullanıcı dosya seçtiyse, dosyayı indir
        if file_path:
            with open(file_path, 'wb') as f:
                f.write(response.content)

            messagebox.showinfo("Güncelleme", "Güncelleme dosyası başarıyla indirildi. Önceki uygulamayı kaldırabilirsiniz. ")
            return file_path
        else:
            messagebox.showinfo("Bilgi", "Dosya seçilmediği için güncelleme işlemi iptal edildi.")
            return None

    except Exception as e:
        messagebox.showerror("Hata", "Güncelleme dosyası indirilemedi: {}".format(str(e)))
        return None

def update_application():
    remote_version = check_for_update()
    if remote_version:
        update_file_path = download_update(remote_version)
        if update_file_path:
            try:
                # Güncelleme dosyasını çalıştır
                subprocess.Popen(update_file_path)

                # Uygulamayı kapat
                sys.exit()

            except Exception as e:
                messagebox.showerror("Hata", "Güncelleme işlemi sırasında bir hata oluştu: {}".format(str(e)))

# Güncelleme butonuna basıldığında bu fonksiyon çağrılacak
def update_button_clicked():
    update_application()
#--------
# .txt dosyasındaki şifreyi kontrol etme
if predefined_password == file_content:
    print("Şifre doğru, uygulama açılıyor...")
    # Uygulamanın açılması için gerekli kodu buraya yazın
else:
    print("Şifre yanlış, uygulama açılmıyor.")
    # Şifre yanlışsa kullanıcıya ikinci bir şifre girmesini sağlayan bir giriş kutusu göster
    second_password = simpledialog.askstring("Şifre Girin", "Güncelleme dosyasını indirmek için geçerli bir şifre girin:")
    # İkinci şifre doğruysa download_update fonksiyonunu çağır
    if second_password == file_content:
        update_application()
    else:
        messagebox.showerror("Hata", "Girilen şifre yanlış. Güncelleme dosyasını indirebilmek için geçerli bir şifre giriniz.")
    # Uygulamadan çık
    sys.exit()
#--------
def create_textbox():
    global frame_count, textboxes

    frame_count += 1
    frame_id = f"frame_{frame_count}"

    frame = Frame(root, bd=1, relief=SOLID, bg="black")
    frame.place(x=1360, y=1080)  # Frame'leri dikey olarak yerleştirir

    textbox = Text(frame, wrap=WORD, height=1, width=25, font=("Helvetica", 12))
    textbox.pack(expand=True, fill=BOTH)  # Text kutusunu doldurur

    # Text kutusuna metni eklerken ilk 3 karakterin sağa hizalanması
    initial_text = " " * 3  # İlk üç karakterlik boşluk
    textbox.insert(END, initial_text)

    # Yazı alanına bir KeyRelease olayı bağlayarak, metnin her değiştirildiğinde başındaki boşluğu yeniden ekleme
    textbox.bind("<KeyRelease>", lambda event, tb=textbox: insert_initial_text(event, tb))

    def insert_initial_text(event, textbox):
        current_text = textbox.get("1.0", "1.3")  # Metnin ilk üç karakterini al
        if current_text.strip() != "   ":  # Eğer metnin ilk üç karakteri sadece boşluktan oluşmuyorsa
            textbox.delete("1.0", "1.3")  # Metnin ilk üç karakterini sil
            textbox.insert("1.0", "   ")  # Yeni başlangıç boşluğunu ekle
    textboxes[frame_id] = (frame, textbox)

    # Sağ üst köşeye bir tutamaç ekleyelim
    handle_upper_right = Frame(frame, bg="black", width=10, height=10, cursor="fleur")
    handle_upper_right.place(relx=1.0, rely=0.0, anchor=NE)

    # Sol üst köşeye bir silme tutamaçı ekleyelim
    handle_upper_left = Frame(frame, bg="black", width=10, height=10, cursor="hand2")
    handle_upper_left.place(x=0, y=0, anchor=NW)  # Tutamaç frame'in sol üst köşesine yapışık olsun
    label_x = Label(handle_upper_left, text="X", fg="black", font=("Arial", 10))
    label_x.pack()

    # Sürükleme işlevselliğini ekleyelim
    handle_upper_right.bind("<Button-1>", lambda event, f=frame: start_drag(event, f))
    handle_upper_right.bind("<B1-Motion>", lambda event, f=frame: drag_frame(event, f))
    handle_upper_right.bind("<ButtonRelease-1>", lambda event, f=frame: stop_drag(event, f))

    # TextBox'ı silme işlevselliğini ekleyelim
    label_x.bind("<Button-1>", lambda event, f=textbox: delete_textbox(event, f))
# Yazı alanında yazıldığında boyutları ayarla
    textbox.bind("<Key>", lambda event, f=textbox: adjust_size(event, f))
def adjust_size(event, textbox):
    # Satırları al
    lines = textbox.get("1.0", "end-1c").split("\n")

    # Satır sayısını ve en uzun satırı hesapla
    num_lines = len(lines)
    longest_line_length = max(len(line) for line in lines)

    # Yazı alanını boyutlandır
    textbox.config(width=longest_line_length + 3, height=num_lines + 1)
def start_drag(event, frame):
    frame.drag_data = {"x": event.x, "y": event.y}


def drag_frame(event, frame):
    x, y = frame.winfo_x() + event.x - frame.drag_data["x"], frame.winfo_y() + event.y - frame.drag_data["y"]
    frame.place(x=x, y=y)


def stop_drag(event, frame):
    frame.drag_data = {}


def delete_textbox(event, textbox):
    textbox.master.destroy()  # TextBox'ın içinde bulunduğu frame'i yok et
    for key, (f, _) in textboxes.items():
        if f == textbox.master:
            del textboxes[key]
            break




def zoom_in():
    global image_img
    scale_factor = 1.1
    if image_img:
        current_width = image_img.width()
        current_height = image_img.height()
        new_width = int(current_width * scale_factor)
        new_height = int(current_height * scale_factor)
        img = Image.open(image_path)
        img = img.resize((new_width, new_height), Image.LANCZOS)
        image_img = ImageTk.PhotoImage(img)
        imageFrame.itemconfig(image_item, image=image_img)

def zoom_out():
    global image_img
    scale_factor = 1.1
    if image_img:
        current_width = image_img.width()
        current_height = image_img.height()
        new_width = int(current_width / scale_factor)
        new_height = int(current_height / scale_factor)
        img = Image.open(image_path)
        img = img.resize((new_width, new_height), Image.LANCZOS)
        image_img = ImageTk.PhotoImage(img)
        imageFrame.itemconfig(image_item, image=image_img)

def second_zoom_in():
    global second_image_img
    scale_factor = 1.1
    if second_image_img:
        current_width = second_image_img.width()
        current_height = second_image_img.height()
        new_width = int(current_width * scale_factor)
        new_height = int(current_height * scale_factor)
        img = Image.open(second_image_path)
        img = img.resize((new_width, new_height), Image.LANCZOS)
        second_image_img = ImageTk.PhotoImage(img)
        imageFrame.itemconfig(second_image_item, image=second_image_img)

def second_zoom_out():
    global second_image_img
    scale_factor = 1.1
    if second_image_img:
        current_width = second_image_img.width()
        current_height = second_image_img.height()
        new_width = int(current_width / scale_factor)
        new_height = int(current_height / scale_factor)
        img = Image.open(second_image_path)
        img = img.resize((new_width, new_height), Image.LANCZOS)
        second_image_img = ImageTk.PhotoImage(img)
        imageFrame.itemconfig(second_image_item, image=second_image_img)

def third_zoom_in():
    global third_image_img
    scale_factor = 1.1
    if third_image_img:
        current_width = third_image_img.width()
        current_height = third_image_img.height()
        new_width = int(current_width * scale_factor)
        new_height = int(current_height * scale_factor)
        img = Image.open(third_image_path)
        img = img.resize((new_width, new_height), Image.LANCZOS)
        third_image_img = ImageTk.PhotoImage(img)
        imageFrame.itemconfig(third_image_item, image=third_image_img)

def third_zoom_out():
    global third_image_img
    scale_factor = 1.1
    if third_image_img:
        current_width = third_image_img.width()
        current_height = third_image_img.height()
        new_width = int(current_width / scale_factor)
        new_height = int(current_height / scale_factor)
        img = Image.open(third_image_path)
        img = img.resize((new_width, new_height), Image.LANCZOS)
        third_image_img = ImageTk.PhotoImage(img)
        imageFrame.itemconfig(third_image_item, image=third_image_img)



def open_image():
    global image_path, image_selected
    desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    image_path = filedialog.askopenfilename(initialdir=desktop_path, title="Resim Seç",
                                                   filetypes=(("JPEG files", "*.jpg"), ("PNG files", "*.png"),
                                                              ("All files", "*.*")))
    if image_path:
        print("Seçilen ikinci dosya yolu:", image_path)
        load_image()
        # Resim seçildiğinde image_selected değişkenini True olarak ayarla
        image_selected = True
    else:
        print("Birinci dosya seçilmedi.")




def load_image():
    global image_path, image_item, image_img
    if image_path:
        print("Resim yükleniyor:", image_path)
        try:
            img = Image.open(image_path)
            canvas_width = imageFrame.winfo_width()
            canvas_height = imageFrame.winfo_height()
            width_ratio = canvas_width / img.width
            height_ratio = canvas_height / img.height
            resize_ratio = min(width_ratio, height_ratio)
            new_width = int(img.width * resize_ratio)
            new_height = int(img.height * resize_ratio)
            img = img.resize((new_width, new_height))
            image_img = ImageTk.PhotoImage(img)
            image_item = imageFrame.create_image((canvas_width - new_width) / 2, (canvas_height - new_height) / 2,
                                                 anchor=NW, image=image_img)
            drag_and_drop(image_item)
            print("Resim yüklendi.")

        except Exception as e:
            print("Resim yüklenirken bir hata oluştu:", e)


def drag_and_drop(image_item):
    start_x, start_y = 0, 0
    dragging = False

    def start_drag(event):
        nonlocal start_x, start_y, dragging
        start_x = event.x
        start_y = event.y
        dragging = True

    def stop_drag(event):
        nonlocal dragging
        dragging = False

    def on_drag(event):
        nonlocal start_x, start_y, dragging
        if dragging:
            imageFrame.move(image_item, event.x - start_x, event.y - start_y)
            start_x = event.x
            start_y = event.y

    imageFrame.tag_bind(image_item, "<Button-1>", start_drag)
    imageFrame.tag_bind(image_item, "<ButtonRelease-1>", stop_drag)
    imageFrame.tag_bind(image_item, "<B1-Motion>", on_drag)


def open_second_image():
    global second_image_path
    desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    second_image_path = filedialog.askopenfilename(initialdir=desktop_path, title="İkinci Resim Seç",
                                                   filetypes=(("JPEG files", "*.jpg"), ("PNG files", "*.png"),
                                                              ("All files", "*.*")))
    if second_image_path:
        print("Seçilen ikinci dosya yolu:", second_image_path)
        load_second_image()
    else:
        print("İkinci dosya seçilmedi.")



def load_second_image():
    global second_image_path, second_image_item, second_image_img
    if second_image_path:
        print("İkinci resim yükleniyor:", second_image_path)
        try:
            img = Image.open(second_image_path)
            canvas_width = imageFrame.winfo_width()
            canvas_height = imageFrame.winfo_height()
            width_ratio = canvas_width / img.width
            height_ratio = canvas_height / img.height
            resize_ratio = min(width_ratio, height_ratio)
            new_width = int(img.width * resize_ratio)
            new_height = int(img.height * resize_ratio)
            img = img.resize((new_width, new_height))
            second_image_img = ImageTk.PhotoImage(img)
            second_image_item = imageFrame.create_image((canvas_width - new_width) / 2,
                                                        (canvas_height - new_height) / 2,
                                                        anchor=NW, image=second_image_img)
            drag_and_drop_second(second_image_item)
            print("İkinci resim yüklendi.")



        except Exception as e:
            print("İkinci resim yüklenirken bir hata oluştu:", e)


def drag_and_drop_second(image_item):
    start_x, start_y = 0, 0
    dragging = False

    def start_drag(event):
        nonlocal start_x, start_y, dragging
        start_x = event.x
        start_y = event.y
        dragging = True

    def stop_drag(event):
        nonlocal dragging
        dragging = False

    def on_drag(event):
        nonlocal start_x, start_y, dragging
        if dragging:
            imageFrame.move(image_item, event.x - start_x, event.y - start_y)
            start_x = event.x
            start_y = event.y

    imageFrame.tag_bind(image_item, "<Button-1>", start_drag)
    imageFrame.tag_bind(image_item, "<ButtonRelease-1>", stop_drag)
    imageFrame.tag_bind(image_item, "<B1-Motion>", on_drag)

def open_third_image():
    global third_image_path
    desktop_path = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    third_image_path = filedialog.askopenfilename(initialdir=desktop_path, title="Üçüncü Resim Seç",
                                                  filetypes=(("JPEG files", "*.jpg"), ("PNG files", "*.png"),
                                                             ("All files", "*.*")))
    if third_image_path:
        print("Seçilen üçüncü dosya yolu:", third_image_path)
        load_third_image()
    else:
        print("Üçüncü dosya seçilmedi.")




def load_third_image():
    global third_image_path, third_image_item, third_image_img
    if third_image_path:
        print("Üçüncü resim yükleniyor:", third_image_path)
        try:
            img = Image.open(third_image_path)
            canvas_width = imageFrame.winfo_width()
            canvas_height = imageFrame.winfo_height()
            width_ratio = canvas_width / img.width
            height_ratio = canvas_height / img.height
            resize_ratio = min(width_ratio, height_ratio)
            new_width = int(img.width * resize_ratio)
            new_height = int(img.height * resize_ratio)
            img = img.resize((new_width, new_height))
            third_image_img = ImageTk.PhotoImage(img)
            third_image_item = imageFrame.create_image((canvas_width - new_width) / 2,
                                                       (canvas_height - new_height) / 2,
                                                       anchor=NW, image=third_image_img)
            drag_and_drop_third(third_image_item)
            print("Üçüncü resim yüklendi.")

        except Exception as e:
            print("Üçüncü resim yüklenirken bir hata oluştu:", e)


def drag_and_drop_third(image_item):
    start_x, start_y = 0, 0
    dragging = False

    def start_drag(event):
        nonlocal start_x, start_y, dragging
        start_x = event.x
        start_y = event.y
        dragging = True

    def stop_drag(event):
        nonlocal dragging
        dragging = False

    def on_drag(event):
        nonlocal start_x, start_y, dragging
        if dragging:
            imageFrame.move(image_item, event.x - start_x, event.y - start_y)
            start_x = event.x
            start_y = event.y

    imageFrame.tag_bind(image_item, "<Button-1>", start_drag)
    imageFrame.tag_bind(image_item, "<ButtonRelease-1>", stop_drag)
    imageFrame.tag_bind(image_item, "<B1-Motion>", on_drag)


def clean_filename(filename):
    return re.sub(r'[^\w\-_. ]', '', filename)






# Resim seçildiğinde bu fonksiyon çağrılır ve image_path güncellenir
def select_image():
    global image_path, image_selected
    # Resim seçimi işlemi
    # image_path değişkeni güncellenir
    # image_selected değişkeni True olarak ayarlanır



#--------------------------EXCEL SAVE
def Excel_Save():
    global image_path, image_item, image_img, image_selected

    M1 = musteriEntry.get('1.0', 'end').strip()
    T1 = tasEntry.get('1.0', 'end').strip()
    K1 = kumasEntry.get('1.0', 'end').strip()
    I1 = iplikEntry.get('1.0', 'end').strip()
    E1 = etiketEntry.get('1.0', 'end').strip()
    A2 = aksesuarEntry.get('1.0', 'end').strip()
    B1 = baskiEntry.get('1.0', 'end').strip()
    N1 = notEntry.get('1.0', 'end').strip()
    O1 = orderEntry.get('1.0', 'end').strip()
    A1 = artikelEntry.get('1.0', 'end').strip()
    D1 = desenEntry.get('1.0', 'end').strip()

    if O1 == "" or A1 == "" or D1 == "":
        messagebox.showerror("Hata", "Lütfen gerekli bilgileri giriniz!")
        return

    # Resim seçilmediyse
    if not image_selected:
        messagebox.showerror("Hata", "Resim seçilmedi!")
        return

    # Excel dosyasını kontrol et ve varsa yükle, yoksa yeni dosya oluştur
    if os.path.exists("İmalatArtwork.xlsx"):
        file = load_workbook("İmalatArtwork.xlsx")
    else:
        file = Workbook()

    sheet = file.active

    # Artikel ve Desen'in daha önce kaydedilip kaydedilmediğini kontrol et
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=4):
        if row[0].value == A1 and row[1].value == D1:
            messagebox.showerror("Hata", "Bu kayıt önceden girilmiş!")
            return

    # İlk satırı atla
    row_index = sheet.max_row + 1

    # Boş satırı bul
    empty_row_index = None
    for row_index in range(2, sheet.max_row + 1):
        if all([sheet.cell(row=row_index, column=col).value is None for col in range(1, sheet.max_column + 1)]):
            empty_row_index = row_index
            break

    # Eğer boş satır bulunamazsa, sonraki satıra yaz
    if empty_row_index is None:
        empty_row_index = sheet.max_row + 1

    # Hücrelere değerleri yazma
    row = [M1, O1, A1, D1, T1, K1, I1, E1, A2, B1, N1]
    # Hücrelere değerleri yazma ve ortalamayı ayarlama
    for index, value in enumerate(row):
        cell = sheet.cell(row=empty_row_index, column=index + 1)
        cell.value = value
        # Hücrenin var olan hizalamasını koru
        cell.alignment = Alignment(wrapText=True)
        # Satır yüksekliğini otomatik olarak ayarla
        sheet.row_dimensions[empty_row_index].auto_size = True

    # Hücre formatlarını ayarlama (wrap text)
    for cell in sheet[empty_row_index]:
        cell.alignment = Alignment(wrapText=True)

    file.save("İmalatArtwork.xlsx")

    # Resmi kaydetme
    img = Image.open(image_path)

    # Geçici olarak küçük bir sürüm oluştur
    temp_img = img.copy()
    temp_img.thumbnail((800, 800))  # İstediğiniz boyuta ayarlayın
    temp_img.save("temp_image.jpg")

    save_filename = (str(D1) + " " + str(A1)).replace("\n", "") + ".jpg"
    save_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "Artwork", save_filename)

    # Resmin boyutunu kontrol et ve belirli boyut sınırına ulaşana kadar kaliteyi azalt
    max_file_size = 1024 * 1024  # 1 MB boyut sınırı (byte cinsinden)
    quality = 150  # Başlangıçta kullanılacak JPEG kalite ayarı (0 ile 100 arasında bir değer)
    temp_img.save(save_path, optimize=True, quality=quality)

    while os.path.getsize(save_path) > max_file_size:
        quality -= 5
        temp_img.save(save_path, optimize=True, quality=quality)

    # Geçici dosyayı sil
    os.remove("temp_image.jpg")

    messagebox.showinfo("Bilgi", "Kayıt başarıyla oluşturuldu!")
#--------------------------EXCEL SAVE

#--------------------------CLEAR
def Clear():
    musteriEntry.delete('1.0', 'end')
    tasEntry.delete('1.0', 'end')
    kumasEntry.delete('1.0', 'end')
    iplikEntry.delete('1.0', 'end')
    etiketEntry.delete('1.0', 'end')
    aksesuarEntry.delete('1.0', 'end')
    baskiEntry.delete('1.0', 'end')
    notEntry.delete('1.0', 'end')
    orderEntry.delete('1.0', 'end')
    artikelEntry.delete('1.0', 'end')
    desenEntry.delete('1.0', 'end')
    imageFrame.delete(image_item)
    delete_image()
#--------------------------CLEAR
#--------------------------SEARCH
searched_artikel = ""  # Global değişken olarak tanımlıyoruz
searched_desen = ""  # Global değişken olarak tanımlıyoruz
original_artikel = ""  # Global değişken olarak tanımlıyoruz
original_desen = ""  # Global değişken olarak tanımlıyoruz
#--------------------------SEARCH

def Search():
    global searched_artikel, searched_desen  # searched_artikel ve searched_desen'i global olarak kullanmak için tanımlıyoruz
    search_desen = str(str(SearchDesen.get()).lower())  # Küçük harfe dönüştürme
    search_artikel = str(str(SearchArtikel.get()).lower())  # Küçük harfe dönüştürme
    Clear()

    if not search_desen or not search_artikel:
        messagebox.showerror("Hata", "Lütfen aramak için desen ve artikel bilgilerini girin!")
        return

    # Excel dosyasını aç
    if os.path.exists("İmalatArtwork.xlsx"):
        file = load_workbook("İmalatArtwork.xlsx")
        sheet = file.active

        # Excel dosyasında arama yap
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=11):
            # Verileri küçük harfe dönüştürerek karşılaştırma yapma
            if str(str(row[2].value).lower()) == str(search_artikel) and str(str(row[3].value).lower()) == str(search_desen):
                searched_artikel = str(row[2].value)  # Bulunan artikel numarasını searched_artikel'e atıyoruz
                searched_desen = str(row[3].value)  # Bulunan desen numarasını searched_desen'e atıyoruz

                # Eşleşen satırı bulduk, bilgileri ilgili alanlara yaz
                musteriEntry.insert('1.0', str(row[0].value) if row[0].value is not None else "")
                tasEntry.insert('1.0', str(row[4].value) if row[4].value is not None else "")
                kumasEntry.insert('1.0', str(row[5].value) if row[5].value is not None else "")
                iplikEntry.insert('1.0', str(row[6].value) if row[6].value is not None else "")
                etiketEntry.insert('1.0', str(row[7].value) if row[7].value is not None else "")
                aksesuarEntry.insert('1.0', str(row[8].value) if row[8].value is not None else "")
                baskiEntry.insert('1.0', str(row[9].value) if row[9].value is not None else "")
                notEntry.insert('1.0', str(row[10].value) if row[10].value is not None else "")
                orderEntry.insert('1.0', str(row[1].value) if row[1].value is not None else "")
                artikelEntry.insert('1.0', str(row[2].value) if row[2].value is not None else "")
                desenEntry.insert('1.0', str(row[3].value) if row[3].value is not None else "")

                # Resmi otomatik olarak çağır
                save_filename = (str(row[3].value) + " " + str(row[2].value)).replace("\n", "") + ".jpg"
                save_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "Artwork", save_filename)

                if os.path.exists(save_path):
                    try:
                        load_image_auto(save_path)
                        print("Resim yüklendi.")
                    except Exception as e:
                        print("Resim yüklenirken bir hata oluştu:", e)
                else:
                    messagebox.showinfo("Bilgi", "Resim bulunamadı!")

                break
        else:
            messagebox.showinfo("Bilgi", "Eşleşen kayıt bulunamadı!")


def load_image_auto(image_path):
    global image_item, image_img
    if image_path:
        print("Resim yükleniyor:", image_path)
        try:
            img = Image.open(image_path)
            canvas_width = imageFrame.winfo_width()
            canvas_height = imageFrame.winfo_height()
            width_ratio = canvas_width / img.width
            height_ratio = canvas_height / img.height
            resize_ratio = min(width_ratio, height_ratio)
            new_width = int(img.width * resize_ratio)
            new_height = int(img.height * resize_ratio)
            img = img.resize((new_width, new_height))
            image_img = ImageTk.PhotoImage(img)
            image_item = imageFrame.create_image((canvas_width - new_width) / 2, (canvas_height - new_height) / 2,
                                                 anchor=NW, image=image_img)
            drag_and_drop(image_item)
        except Exception as e:
            print("Resim yüklenirken bir hata oluştu:", e)




#--------------------------SEARCH
#--------------------------UPDATE


def overwrite():
    global searched_artikel, searched_desen  # searched_artikel ve searched_desen'i global olarak kullanmak için tanımlıyoruz

    # Kullanıcı tarafından girilen yeni verileri alın
    M1 = musteriEntry.get('1.0', 'end').strip() or None
    T1 = tasEntry.get('1.0', 'end').strip() or None
    K1 = kumasEntry.get('1.0', 'end').strip() or None
    I1 = iplikEntry.get('1.0', 'end').strip() or None
    E1 = etiketEntry.get('1.0', 'end').strip() or None
    A2 = aksesuarEntry.get('1.0', 'end').strip() or None
    B1 = baskiEntry.get('1.0', 'end').strip() or None
    N1 = notEntry.get('1.0', 'end').strip() or None
    O1 = orderEntry.get('1.0', 'end').strip() or None
    A1 = artikelEntry.get('1.0', 'end').strip() or None
    D1 = desenEntry.get('1.0', 'end').strip() or None

    # Eğer değişiklik yapılan kayıtın artikel veya desen numarası değiştirilmişse uyarı ver ve işlemi iptal et
    if searched_artikel != A1 or searched_desen != D1:
        messagebox.showwarning("Uyarı", "Artikel veya Desen Numaralarını Değiştiremezsiniz!")
        # Desen ve artikel giriş alanlarının içeriğini güncelle
        desenEntry.delete("1.0", "end")  # Önce mevcut içeriği sil
        desenEntry.insert("1.0", str(SearchDesen.get()).strip())  # Yeni değeri ekle

        artikelEntry.delete("1.0", "end")  # Önce mevcut içeriği sil
        artikelEntry.insert("1.0", str(SearchArtikel.get()).strip())  # Yeni değeri ekle

        return

    # Varolan Excel dosyasını yükleyin
    file = load_workbook("İmalatArtwork.xlsx")
    sheet = file.active

    # Belirli bir şartı sağlayan satırı bulma
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        if row[2].value == A1 and row[3].value == D1:  # Artikel numarası ve desen numarası aynı olan satırı bul
            # Hücrelerin değerlerini güncelle
            row[0].value = M1
            row[1].value = O1
            row[4].value = T1
            row[5].value = K1
            row[6].value = I1
            row[7].value = E1
            row[8].value = A2
            row[9].value = B1
            row[10].value = N1
            break  # İlk eşleşmeyi bulduktan sonra döngüden çık

    # Değişiklikleri kaydedin ve Excel dosyasını kapatın
    file.save("İmalatArtwork.xlsx")
    # Resmi kaydetme
    img = Image.open(image_path)

    # Geçici olarak küçük bir sürüm oluştur
    temp_img = img.copy()
    temp_img.thumbnail((800, 800))  # İstediğiniz boyuta ayarlayın
    temp_img.save("temp_image.jpg")

    save_filename = (str(D1) + " " + str(A1)).replace("\n", "") + ".jpg"
    save_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "Artwork", save_filename)

    # Resmin boyutunu kontrol et ve belirli boyut sınırına ulaşana kadar kaliteyi azalt
    max_file_size = 1024 * 1024  # 1 MB boyut sınırı (byte cinsinden)
    quality = 150  # Başlangıçta kullanılacak JPEG kalite ayarı (0 ile 100 arasında bir değer)
    temp_img.save(save_path, optimize=True, quality=quality)

    while os.path.getsize(save_path) > max_file_size:
        quality -= 5
        temp_img.save(save_path, optimize=True, quality=quality)

    # Geçici dosyayı sil
    os.remove("temp_image.jpg")

    messagebox.showinfo("Bilgi", "Kayıt Güncellendi!")


def delete_image():
    global image_path, second_image_path, third_image_path, image_selected,image_img
    image_img = ""
    image_path = ""
    second_image_path = ""
    third_image_path = ""
    # Eğer birinci resim yüklenmişse, onu sil
    if image_path:
        imageFrame.delete(image_item)
        image_path = ""
        print("Birinci resim başarıyla silindi.")

    # Eğer ikinci resim yüklenmişse, onu sil
    if second_image_path:
        imageFrame.delete(second_image_item)
        second_image_path = ""
        print("İkinci resim başarıyla silindi.")

    # Eğer üçüncü resim yüklenmişse, onu sil
    if third_image_path:
        imageFrame.delete(third_image_item)
        third_image_path = ""
        print("Üçüncü resim başarıyla silindi.")

    image_selected = False


def take_screenshot():
    global root, footer_frame, outer_frame
    # Get current date and time
    now = datetime.now()
    timestamp = now.strftime("%d-%m-%Y %H-%M-%S")
    # Create file name based on Label texts
    file_name = f"{desenEntry.get('1.0', 'end').strip()} {artikelEntry.get('1.0', 'end').strip()} {musteriEntry.get('1.0', 'end').strip()} {orderEntry.get('1.0', 'end').strip()} {timestamp}.jpg"
    # Ask user for save location
    save_path = filedialog.asksaveasfilename(defaultextension=".jpg",
                                             filetypes=(("JPEG files", "*.jpg"), ("All files", "*.*")),
                                             initialfile=file_name)
    if save_path:
        try:
            # Update the root window to ensure all changes are rendered
            root.update()
            # Get the coordinates of the footer frame relative to the screen
            x_footer, y_footer = footer_frame.winfo_rootx(), footer_frame.winfo_rooty()
            # Get the coordinates of the outer frame relative to the screen
            x_outer, y_outer = outer_frame.winfo_rootx(), outer_frame.winfo_rooty()
            # Take screenshot of the root window excluding the outer and footer frames
            im = ImageGrab.grab(bbox=(root.winfo_rootx(), y_outer + outer_frame.winfo_height(),
                                       root.winfo_rootx() + root.winfo_width(),
                                       y_footer))
            # Save the screenshot
            im.save(save_path, quality=95)
            print("Screenshot başarıyla kaydedildi:", save_path)
        except Exception as e:
            print("Screenshot alınırken bir hata oluştu:", e)
    else:
        print("Kaydedilecek konum seçilmedi.")








root = CTk()
root.title("İmalat Artwork")
root.geometry("1600x1200")
customtkinter.set_appearance_mode("dark")
root.config(bg="white")  # Arkaplan rengini beyaz yapar
root.iconbitmap(resource_path('missy_2.ico'))
root.resizable(width=False, height=False)  # Pencerenin boyutunu sabit tutar
# Pencereyi ekranın tam ortasına konumlandırma
window_width = 1600  # Pencere genişliği
window_height = 1200  # Pencere yüksekliği
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = (screen_width - window_width) // 2
y = (screen_height - window_height) // 2
root.geometry("+{}+{}".format(x, y))
file=pathlib.Path("İmalatArtwork.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"] = "Müşteri"
    sheet["B1"] = "Order"
    sheet["C1"] = "Artikel"
    sheet["D1"] = "Desen"
    sheet["E1"] = "Taş"
    sheet["F1"] = "Kumaş"
    sheet["G1"] = "İplik"
    sheet["H1"] = "Etiket"
    sheet["I1"] = "Aksesuar"
    sheet["J1"] = "Baskı"
    sheet["K1"] = "Not"

    file.save("İmalatArtwork.xlsx")






# Dış çerçeve
outer_frame = Frame(root, bg="white")
outer_frame.pack(fill=X)

# Ortadaki çerçeve
center_frame = Frame(outer_frame, bg="white")
center_frame.pack(side=LEFT, expand=True)



headingLabel = Label(center_frame, text="İmalat Artwork".upper(),
                     font=("times new roman", 30, "bold"),
                     bg="white", fg="black", bd=12, relief=FLAT)
headingLabel.pack()
headingLabel.pack(padx=10, pady=10)

#arama kutusu
#Başlık etiketi oluşturma
title_label = str(Label(root, text="Artikel",bg="white", font="arial 15").place(x=25,y=10))
SearchArtikel=StringVar()
Entry(outer_frame,textvariable=SearchArtikel,width=12,bd=2,font="arial 20").place(x=25,y=40)

title_label = str(Label(root, text="Desen",bg="white", font="arial 15").place(x=265,y=10))
SearchDesen=StringVar()
Entry(outer_frame,textvariable=SearchDesen,width=12,bd=2,font="arial 20").place(x=235,y=40)

# Arama Butonu
Excel_button = CTkButton(root, text="\u2315",width=15,height=37, font=("Arial",20),bg_color="white",fg_color=("#e97373"), hover_color="black",command=Search)
Excel_button.place(x=425,y=40)  # Koordinatları değiştirerek butonun yerini belirleyin.

# Overwrite butonunu oluşturma ve fonksiyonu bağlama
Overwrite_button = CTkButton(root, text="\u21BB",width=15,height=37, font=("Arial",20),bg_color="white",fg_color=("#e97373"), hover_color="black",command=overwrite)
Overwrite_button.place(x=455,y=40)  # Koordinatları değiştirerek butonun yerini belirleyin.

# Excel Kaydetme Butonu
Excel_button = CTkButton(root, text="\u2713",width=15,height=37, font=("Arial",20),bg_color="white",fg_color=("#e97373"), hover_color="black",command=Excel_Save)
Excel_button.place(x=489,y=40)  # Koordinatları değiştirerek butonun yerini belirleyin.

# Sağdaki çerçeve
right_frame = Frame(outer_frame, bg="white", width=200)
right_frame.pack(side=RIGHT)

# İkon
icon_image = Image.open(resource_path('missy.png'))
new_icon_size = (100, 60)
resized_icon_image = icon_image.resize(new_icon_size)
icon_photo = ImageTk.PhotoImage(resized_icon_image)
icon_label = Label(right_frame, image=icon_photo, bg="white")
icon_label.pack(padx=10, pady=0, anchor="center")


# Define image path
image_path = ""
mainFrame = Frame(root, relief=RIDGE, bd=2, bg="white")
mainFrame.pack(pady=5)

textFrame = LabelFrame(mainFrame, bg="white", height=1008, relief=FLAT, bd=4)
textFrame.grid(row=0, column=0)
# -----------------MUSTERI
musteriLabel = Label(textFrame, text="Müşteri:", font=("times new roman", 15, "bold"),
                     bg="white", fg="black", bd=3)
musteriLabel.grid(row=0, column=0)

musteriEntry = Text(textFrame, font=("times new roman", 15, "bold"),
                    bg="white", fg="black", relief=RIDGE, bd=2, height=1, width=30)
musteriEntry.grid(row=0, column=1, padx=0, pady=5)
# -----------------MUSTERI
# -----------------ORDER
orderLabel = Label(textFrame, text="Order:", font=("times new roman", 15, "bold"),
                   bg="white", fg="black", bd=3)
orderLabel.grid(row=1, column=0)

orderEntry = Text(textFrame, font=("times new roman", 15, "bold"),
                  bg="white", fg="black", relief=RIDGE, bd=2, height=1, width=30)
orderEntry.grid(row=1, column=1, padx=0, pady=5)
# -----------------ORDER
# -----------------ARTİKEL
artikelLabel = Label(textFrame, text="Artikel:", font=("times new roman", 15, "bold"),
                     bg="white", fg="black", bd=3)
artikelLabel.grid(row=2, column=0)

artikelEntry = Text(textFrame, font=("times new roman", 15, "bold"),
                    bg="white", fg="black", relief=RIDGE, bd=2, height=1, width=30)
artikelEntry.grid(row=2, column=1, padx=0, pady=5)
# -----------------ARTİKEL
# -----------------DESEN
desenLabel = Label(textFrame, text="Desen:", font=("times new roman", 15, "bold"),
                   bg="white", fg="black", bd=3)
desenLabel.grid(row=3, column=0)

desenEntry = Text(textFrame, font=("times new roman", 15, "bold"),
                  bg="white", fg="black", relief=RIDGE, bd=2, height=1, width=30)
desenEntry.grid(row=3, column=1, padx=0, pady=5)
# -----------------DESEN
# -----------------TAŞ
tasLabel = Label(textFrame, text="Taş:", font=("times new roman", 15, "bold"),
                   bg="white", fg="black", bd=3)
tasLabel.grid(row=4, column=0)

tasEntry = Text(textFrame, font=("times new roman", 15, "bold"),
                  bg="white", fg="black", relief=RIDGE, bd=2, height=1, width=30)
tasEntry.grid(row=4, column=1, padx=0, pady=5)
# -----------------TAŞ
# -----------------KUMAŞ
kumasLabel = Label(textFrame, text="Kumaş:", font=("times new roman", 15, "bold"),
                   bg="white", fg="black", bd=3)
kumasLabel.grid(row=5, column=0)

kumasEntry = Text(textFrame, font=("times new roman", 15, "bold"),
                  bg="white", fg="black", relief=RIDGE, bd=2, height=1, width=30)
kumasEntry.grid(row=5, column=1, padx=0, pady=5)
# -----------------KUMAŞ
# -----------------İPLİK
iplikLabel = Label(textFrame, text="İplik:", font=("times new roman", 15, "bold"),
                   bg="white", fg="black", bd=3)
iplikLabel.grid(row=6, column=0)

iplikEntry = Text(textFrame, font=("times new roman", 15, "bold"),
                  bg="white", fg="black", relief=RIDGE, bd=2, height=1, width=30)
iplikEntry.grid(row=6, column=1, padx=0, pady=5)
# -----------------İPLİK
# -----------------ETİKET
etiketLabel = Label(textFrame, text="Etiket:", font=("times new roman", 15, "bold"),
                    bg="white", fg="black", bd=3)
etiketLabel.grid(row=7, column=0)

etiketEntry = Text(textFrame, font=("times new roman", 15, "bold"),
                   bg="white", fg="black", relief=RIDGE, bd=2, height=1, width=30)
etiketEntry.grid(row=7, column=1, padx=0, pady=5)
# -----------------ETİKET
# -----------------AKSESUAR
aksesuarLabel = Label(textFrame, text="Aksesuar:", font=("times new roman", 15, "bold"),
                      bg="white", fg="black", bd=3)
aksesuarLabel.grid(row=8, column=0)

aksesuarEntry = Text(textFrame, font=("times new roman", 15, "bold"),
                     bg="white", fg="black", relief=RIDGE, bd=2, height=6, width=30)
aksesuarEntry.grid(row=8, column=1, padx=0, pady=5)
# -----------------AKSESUAR
# -----------------BASKI
baskiLabel = Label(textFrame, text="Baskı:", font=("times new roman", 15, "bold"),
                   bg="white", fg="black", bd=3)
baskiLabel.grid(row=9, column=0)

baskiEntry = Text(textFrame, font=("times new roman", 15, "bold"),
                  bg="white", fg="black", relief=RIDGE, bd=2, height=6, width=30)
baskiEntry.grid(row=9, column=1, padx=0, pady=5)
# -----------------BASKI
# -----------------NOT

notLabel = Label(textFrame, text="Not:", font=("times new roman", 15, "bold"),
                 bg="white", fg="black", bd=3)
notLabel.grid(row=10, column=0)
notEntry = Text(textFrame, font=("times new roman", 15, "bold"),
                bg="white", fg="black", relief=RIDGE, bd=2, height=16, width=30)
notEntry.grid(row=10, column=1, padx=0, pady=5)
# -----------------NOT


def add_default_text_to_labels():
    default_texts = {
        musteriEntry: "Tunika",
        orderEntry: "",
        artikelEntry: "",
        desenEntry: "",
        kumasEntry: "",
        iplikEntry: "",
        etiketEntry: "",
        aksesuarEntry: "",
        baskiEntry: "Konvör: 198°C 5 HIZ",
        notEntry: ""
    }
    for entry, default_text in default_texts.items():
        if not entry.get("1.0", "end").strip():  # Check if the entry is empty
            entry.insert("1.0", default_text)


drawings = []
brush_size = 2  # Varsayılan fırça boyutu
drawing_mode = ""
brush_color = "black"  # Başlangıçta kullanılacak fırça rengi

def start_line(event):
    global start_x, start_y
    start_x, start_y = event.x, event.y

def draw_line(event):
    global start_x, start_y
    x, y = event.x, event.y
    # Mevcut çizgiyi temizle
    imageFrame.delete("temp_line")
    # Yeni çizgiyi çiz
    imageFrame.create_line(start_x, start_y, x, y, fill=brush_color, width=brush_size, tags=("drawing", "temp_line"))

def stop_line(event):
    global start_x, start_y
    x, y = event.x, event.y
    # Mevcut çizgiyi temizle
    imageFrame.delete("temp_line")
    # Gerçek çizgiyi oluştur
    imageFrame.create_line(start_x, start_y, x, y, fill=brush_color, width=brush_size, tags="drawing")

def start_or_draw(event):
    global last_x, last_y
    x, y = event.x, event.y
    if drawing_mode == "line":
        start_line(event)
    elif drawing_mode == "brush":
        last_x, last_y = x, y

def draw_or_line(event):
    global last_x, last_y
    x, y = event.x, event.y
    if drawing_mode == "line":
        draw_line(event)
    elif drawing_mode == "brush":
        # Burada önceki "draw" fonksiyonunun yerini alacak işlevsellik olmalı
        if last_x is not None and last_y is not None:
            imageFrame.create_line(last_x, last_y, x, y, fill=brush_color, width=brush_size, tags="drawing")
            last_x, last_y = x, y

def stop_or_stop(event):
    global last_x, last_y
    x, y = event.x, event.y
    if drawing_mode == "line":
        stop_line(event)
    elif drawing_mode == "brush":
        pass  # "stop_drawing" fonksiyonu artık gerekli değil


def toggle_brush_mode():
    global drawing_mode
    if drawing_mode == "brush":
        drawing_mode = "line"
        brush_button.configure(text="Fırça")  # Mod değiştiğinde buton metnini "Fırça" olarak güncelle
    else:
        drawing_mode = "brush"
        brush_button.configure(text="Çizgi")  # Mod değiştiğinde buton metnini "Çizgi" olarak güncelle


def select_brush_color():
    global brush_color
    color = colorchooser.askcolor(title="Fırça Rengini Seçin")  # Renk seçim paletini açın
    if color:
        brush_color = color[1]  # Seçilen rengin RGB kodunu alın

def change_brush_size(new_size):
    global brush_size
    brush_size = int(new_size)





def clear_canvas():
    global imageFrame
    if imageFrame.find_withtag("drawing"):  # Eğer çizim varsa
        imageFrame.delete("drawing")  # Çizimi sil
    else:
        print("Silinecek çizim bulunamadı.")





# Create canvas for image
imageFrame = Canvas(mainFrame, bg="white", width=1130, height=992)
imageFrame.grid(row=0, column=1)
# Create canvas for image
BosFrame = Frame(mainFrame, bg="white", width=10, height=985)
BosFrame.grid(row=0, column=2)

# Fare olaylarını bağlayalım
imageFrame.bind("<Button-3>", start_or_draw)
imageFrame.bind("<B3-Motion>", draw_or_line)
imageFrame.bind("<ButtonRelease-3>", stop_or_stop)


# Call the function to add default text to the specified labels
add_default_text_to_labels()

# Footer için bir frame oluşturalım
footer_frame = Frame(root, bg="white", height=10)
footer_frame.pack(side=BOTTOM, fill=X)

# Create select image button
select_image_button = CTkButton(footer_frame, text="Resim", command=open_image, height=50, width=80, font=("impact",25), fg_color=("#e97373"), hover_color="black")
select_image_button.pack(side=LEFT, padx=5, pady=15)
zoom_in_button = CTkButton(footer_frame, text="+", command=zoom_in, height=50, width=80, font=("impact",25), fg_color=("#e97373"), hover_color="black")
zoom_in_button.pack(side=LEFT, padx=5, pady=15)

zoom_out_button = CTkButton(footer_frame, text="-", command=zoom_out, height=50, width=80, font=("impact",25), fg_color=("#e97373"), hover_color="black")
zoom_out_button.pack(side=LEFT, padx=5, pady=15)

# Create second image button
select_second_image_button = CTkButton(footer_frame, text="İkinci Resim", command=open_second_image, height=50, width=80, font=("impact",25), fg_color=("#e97373"), hover_color="black")
select_second_image_button.pack(side=LEFT, padx=5, pady=15)

second_zoom_in_button = CTkButton(footer_frame, text="+", command=second_zoom_in, height=50, width=80, font=("impact",25), fg_color=("#e97373"), hover_color="black")
second_zoom_in_button.pack(side=LEFT, padx=5, pady=15)

second_zoom_out_button = CTkButton(footer_frame, text="-", command=second_zoom_out, height=50, width=80, font=("impact",25), fg_color=("#e97373"), hover_color="black")
second_zoom_out_button.pack(side=LEFT, padx=5, pady=15)

# Create third image button
select_third_image_button = CTkButton(footer_frame, text="Üçüncü Resim", command=open_third_image, height=50, width=80, font=("impact",25), fg_color=("#e97373"), hover_color="black")
select_third_image_button.pack(side=LEFT, padx=5, pady=15)

third_zoom_in_button = CTkButton(footer_frame, text="+", command=third_zoom_in, height=50, width=80, font=("impact",25), fg_color=("#e97373"), hover_color="black")
third_zoom_in_button.pack(side=LEFT, padx=5, pady=15)

third_zoom_out_button = CTkButton(footer_frame, text="-", command=third_zoom_out, height=50, width=80, font=("impact",25), fg_color=("#e97373"), hover_color="black")
third_zoom_out_button.pack(side=LEFT, padx=5, pady=15)
# Create delete image button
delete_image_button = CTkButton(footer_frame, text="TEMİZLE", command=delete_image, height=50, width=80, font=("impact",25), fg_color=("#e97373"), hover_color="black")
delete_image_button.pack(side=LEFT, padx=5, pady=15)



# Fırça butonu
brush_button = CTkButton(footer_frame, text="Fırça", height=50, width=40, font=("impact",25), fg_color=("#e97373"), hover_color="black", command=toggle_brush_mode)
brush_button.pack(side=LEFT, padx=5, pady=15)

# Create clear canvas button
clear_canvas_button = CTkButton(footer_frame, text="Çizimi Sil", command=clear_canvas, height=50, width=40, font=("impact",25), fg_color=("#e97373"), hover_color="black")
clear_canvas_button.pack(side=LEFT, padx=5, pady=15)

brush_size_slider = CTkSlider(footer_frame, from_=1, to=20, command=change_brush_size, width=150, progress_color=("#e97373"), button_color=("black"),button_hover_color=("black"))
brush_size_slider.set(brush_size)  # Varsayılan fırça boyutunu ayarla
brush_size_slider.pack(side=LEFT, padx=5, pady=15)

color_button = CTkButton(footer_frame, text="Renk", command=select_brush_color, height=50, width=40, font=("impact",25), fg_color=("#e97373"), hover_color="black")
color_button.pack(side=LEFT, padx=5, pady=15)  # Renk seçim butonunu ekleyin

create_button = CTkButton(footer_frame, text="Text", command=create_textbox, height=50, width=40, font=("impact",25), fg_color=("#e97373"), hover_color="black")
create_button.pack(side=LEFT, padx=5, pady=15)  # Renk seçim butonunu ekleyin

# Create screenshot button
screenshot_button = CTkButton(footer_frame, text="JPG", command=take_screenshot, height=50, width=40, font=("impact",25), fg_color=("#e97373"), hover_color="black")
screenshot_button.pack(side=LEFT, padx=5, pady=15)

# Güncelleme butonunu oluşturma ve fonksiyonu bağlama
update_button = CTkButton(footer_frame, text="Güncelle", command=update_button_clicked,height=1, width=1, text_color="black",font=("Arial",10), fg_color=("white"), hover_color="white")
update_button.place(x=5, y=65)  # Koordinatları değiştirerek butonun yerini belirleyin.

# Metin etiketi oluştur
version_label = CTkLabel(footer_frame, text="Version: {}".format(LOCAL_VERSION),height=5, width=4, text_color="black")
version_label.place(x=55, y=67)


# Configure mainFrame to center the button
mainFrame.grid_rowconfigure(0, weight=1)
mainFrame.grid_rowconfigure(2, weight=1)
mainFrame.grid_columnconfigure(0, weight=1)
mainFrame.grid_columnconfigure(2, weight=1)

root.mainloop()
